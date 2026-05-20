#!/usr/bin/env python3
"""
Sync the May 2026 Excel cost-side data into the platform:

  1. Insert missing parts referenced by the Filter Cost tab.
  2. Wire engine_filters (engine_type_id → part_id) for the filter-variant
     engines created by the earlier May 2026 sync.
  3. Attach a $8 "Oil Change Labour" service_cost to each engine's primary
     oil filter, so the Labour column matches the Excel "Fuel" $8 entry.
  4. Update oil_types per-litre cost columns from Excel "Oil Price" tab.
  5. Seed volume_tiers (8-20 / 21-38 / 39-46 / 47+) from the per-oil tab
     headers so the Tier+ column populates.

Writes:
  supabase/seed/may2026_filter_cost_and_oils.sql   — apply via psql / SQL editor

Run:
  SUPABASE_SERVICE_ROLE_KEY=... python3 scripts/sync-may-2026-costs.py
"""
from __future__ import annotations

import json
import os
import re
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install with: pip3 install openpyxl")


REPO = Path(__file__).resolve().parents[1]
WORKBOOK = REPO / "doc" / "May 2026 Standard.xlsx"
OUT = REPO / "supabase" / "seed" / "may2026_filter_cost_and_oils.sql"

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "https://facbmuiigusuekddtscz.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_KEY:
    sys.exit("SUPABASE_SERVICE_ROLE_KEY env var is required")


# Re-use the engine mapping from the earlier seed script.
# Maps Excel column-A label → canonical DB display_name.
ENGINE_LABEL_TO_DISPLAY: dict[str, str] = {
    "C 7 / 10 / 3126 With Cat Filter":                  "Cat C7/C10/3126 With Cat Filter",
    "C 7 / 10 / 3126 With Fleetguard Filter":           "Cat C7/C10/3126 With Fleetguard Filter",
    "C 12 / 3406 With  Cat Filter":                     "Cat C12/3406 With Cat Filter",
    "C 12 / 3406 With Cat Filter":                      "Cat C12/3406 With Cat Filter",
    "C 12 / 3406 With Fleetguard Filter":               "Cat C12/3406 With Fleetguard Filter",
    "C 13 / 15 With  Cat Filter":                       "Cat C13/C15 With Cat Filter",
    "C 13 / 15 With Cat Filter":                        "Cat C13/C15 With Cat Filter",
    "C 13 / 15 With Fleetguard Filter":                 "Cat C13/C15 With Fleetguard Filter",
    "Cummins N14":                                      "Cummins N14",
    "Cummins 5.9L/6.7L":                                "Cummins 5.9L/6.7L",
    "Cummins ISC / ISL":                                "Cummins ISC/ISL",
    "Cummins ISB":                                      "Cummins ISB",
    "Cummins ISM":                                      "Cummins ISM",
    "Cummins ISX":                                      "Cummins ISX/X15",
    "Cummins ISX / X15":                                "Cummins ISX/X15",
    "Cummins M11":                                      "Cummins M11",
    "Detroit With Detroit Filter":                      "Detroit 60 Series With Detroit Filter",
    "Detroit With Fleetguard":                          "Detroit 60 Series With Fleetguard Filter",
    "Detroit With Detroit Filter (DD8)":                "Detroit DD8 With Detroit Filter",
    "Detroit With Detroit Filter (DD13,DD15,DD16)":     "Detroit DD13/DD15/DD16 With Detroit Filter",
    "Detroit With Fleetguard Filter (DD13,DD15.DD16)":  "Detroit DD13/DD15/DD16 With Fleetguard Filter",
    "Duramax / Vortec 8100L":                           "Duramax / Vortec 8100L",
    "F250 / F350 / F550 / F650":                        "Ford F250/F350/F550/F650",
    "Hino":                                             "Hino With Hino Filter",
    "Hino With Fleetguard":                             "Hino With Fleetguard Filter",
    "International":                                    "International With Fleetguard Filter",
    "International S13":                                "International S13",
    "International Vista":                              "International Vista",
    "Navistar":                                         "International Navistar",
    "Isuzu":                                            "Isuzu With Isuzu Filter",
    "Isuzu With Fleetguard Filter":                     "Isuzu With Fleetguard Filter",
    "Mack With Mack Filter":                            "Mack With Mack Filter",
    "Mack With Fleetguard Filter":                      "Mack With Fleetguard Filter",
    "Mack With Mack Filter( 52L )":                     "Mack 52L With Mack Filter",
    "Mack With Fleetguard Filter(52L)":                 "Mack 52L With Fleetguard Filter",
    "MP 7 / MP8 With Mack":                             "Mack MP 7/8 With Mack Filter",
    "MP 7 / MP8 With Mack Filter":                      "Mack MP 7/8 With Mack Filter",
    "MP 7 / MP8 With Fleetguard":                       "Mack MP 7/8 With Fleetguard Filter",
    "MP 7 / MP8 With Fleetguard Filter":                "Mack MP 7/8 With Fleetguard Filter",
    "MP 7/MP8/ MP10 With Mack (52L)":                   "Mack MP 7/8/10 52L With Mack Filter",
    "MP 7/MP8/ MP10 With Fleetguard ( 52L)":            "Mack MP 7/8/10 52L With Fleetguard Filter",
    "Max Force 7":                                      "MaxForce 7/9/10 With MaxForce Filter",
    "Max Force 7 / 9 / 10":                             "MaxForce 7/9/10 With MaxForce Filter",
    "Max Force 7 With Fleetguard":                      "MaxForce 7/9/10 With Fleetguard Filter",
    "Max Force 7 / 9 / 10 With Fleetguard":             "MaxForce 7/9/10 With Fleetguard Filter",
    "Max Force 13":                                     "MaxForce 13 With MaxForce Filter",
    "Max Force 13 With Fleetguard":                     "MaxForce 13 With Fleetguard Filter",
    "Mercedes  Benz 4000":                              "Mercedes Benz 4000 With Mercedes Filter",
    "Mercedes  Benz 4000 With Fleeguard Filter":        "Mercedes Benz 4000 With Fleetguard Filter",
    "Mercedes  Benz 900":                               "Mercedes Benz 900 With Mercedes Filter",
    "Mercedes  Benz 900 With Fleetguard Filter":        "Mercedes Benz 900 With Fleetguard Filter",
    "Mercedes Benz Sprinter":                           "Mercedes Benz Sprinter",
    "Mitsubushi":                                       "Mitsubishi Engine",
    "Paccer":                                           "Paccar With Paccar Filter",
    "Paccer With Fleetguard":                           "Paccar With Fleetguard Filter",
    "Paccer Small":                                     "Paccar Small",
    "Volvo With Volvo Filter":                          "Volvo D12 With Volvo Filter",
    "Volvo With Fleetguard":                            "Volvo D12 With Fleetguard Filter",
    "Volvo With Volvo Filter (D16 42L)":                "Volvo D16 42L With Volvo Filter",
    "Volvo With Fleetguard(D16 42L)":                   "Volvo D16 42L With Fleetguard Filter",
    "Volvo With Volvo Filter (D16 52L)":                "Volvo D16 52L With Volvo Filter",
    "Volvo With Fleetguard(D16 52L)":                   "Volvo D16 52L With Fleetguard Filter",
}


# Oil tab → DB oil_types.code (canonical) — same mapping as the earlier sync.
EXCEL_OIL_PRICE: list[tuple[str, str, float, float]] = []  # (db_code, oil_name, bulk_per_L, gallon_per_L)


def sb(p):
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{p}",
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def sql_escape(s: str) -> str:
    return s.replace("'", "''")


# ----------------------------------------------------------------------------
# Part-number → category inference from prefix. Used when inserting a missing
# part — we still need to pick a category_id.
# ----------------------------------------------------------------------------
CATEGORY_PREFIX_RULES: list[tuple[re.Pattern, str]] = [
    (re.compile(r"^LF",  re.I), "Oil Filter"),
    (re.compile(r"^FF",  re.I), "Fuel Filter"),
    (re.compile(r"^FS",  re.I), "Fuel Separetor"),
    (re.compile(r"^FK",  re.I), "Fuel Filter"),
    (re.compile(r"^AF",  re.I), "Air Filter"),
    (re.compile(r"^CA",  re.I), "Air Filter"),
    (re.compile(r"^AH",  re.I), "Air Filter"),
    (re.compile(r"^AD",  re.I), "Air Filter"),
    (re.compile(r"^AC",  re.I), "Air Filter"),
    (re.compile(r"^CS",  re.I), "Coolant Filter"),
    (re.compile(r"^CV",  re.I), "Cab Filter"),
    (re.compile(r"^CF",  re.I), "Cab Filter"),
    (re.compile(r"^1R",  re.I), "Oil Filter"),  # Cat OE oil filter prefix
    (re.compile(r"^P5",  re.I), "Oil Filter"),  # Donaldson oil
    (re.compile(r"^P55", re.I), "Oil Filter"),
    (re.compile(r"^P15", re.I), "Air Filter"),
    (re.compile(r"^P18", re.I), "Air Filter"),
    (re.compile(r"^P55", re.I), "Fuel Filter"),
    (re.compile(r"^B",   re.I), "Oil Filter"),  # Baldwin
    (re.compile(r"^BF",  re.I), "Fuel Filter"),
]


def infer_category(part_number: str) -> str:
    for pat, cat in CATEGORY_PREFIX_RULES:
        if pat.match(part_number):
            return cat
    return "Misc"


# ----------------------------------------------------------------------------
# Parse Excel Filter Cost tab.
# Each engine occupies 2 rows: row N has label + part #s, row N+1 has costs.
# Cols 2-3 = Oil filters; cols 4-5 = Fuel filters.
# ----------------------------------------------------------------------------
def parse_filter_cost(wb):
    sh = wb["Filter Cost"]
    out: list[dict] = []
    r = 6
    while r <= sh.max_row:
        label = sh.cell(r, 1).value
        if not label:
            r += 1
            continue
        label_s = str(label).strip()
        if label_s in ENGINE_LABEL_TO_DISPLAY:
            entries: list[dict] = []
            for c, role in [(2, "oil"), (3, "oil"), (4, "fuel"), (5, "fuel")]:
                pn = sh.cell(r, c).value
                cost = sh.cell(r + 1, c).value
                if pn and isinstance(cost, (int, float)):
                    entries.append({
                        "part_number": str(pn).strip(),
                        "cost": float(cost),
                        "role": role,
                        "col": c,
                    })
            if entries:
                out.append({
                    "excel_label": label_s,
                    "display_name": ENGINE_LABEL_TO_DISPLAY[label_s],
                    "filters": entries,
                })
        r += 2
    return out


# ----------------------------------------------------------------------------
# Parse Excel Oil Price tab.
# Rows alternate: row N = brand/part numbers, row N+1 = name + cost data.
# Cols B-H = Bulk Sell, Bulk Cost, Gallon Case Cost, Gallon Cost (per gallon),
# Gallon Per Litre Cost, Gallon Sell raw, Gallon Sell .99.
# We care about Bulk Cost (col C, per litre) and Gallon Per Litre Cost (col F).
# ----------------------------------------------------------------------------
def parse_oil_price(wb) -> dict[str, dict]:
    sh = wb["Oil Price"]
    # Match by oil name → DB code. The DB names are quite different from
    # Excel labels — use substring matching after normalising.
    EXCEL_LABEL_TO_DB_CODE = {
        "T4  ( 15W40 )":          "500010048",   # Shell T4 15W40
        "T5  ( 10W30 )":          "500010047",   # Shell T5 10W30
        "T6  ( 5W30 ) SYN":       "14",          # Shell T6 5W30
        "T6 5W40 Gallon":         "9",           # Shell T6 5W40
        "15W40 400 XLE SB":       "257004",      # Delo 400 XLE SB 15W40
        "10W30 400 XLE Syn Blend":"10",          # Delo 400LE 10W30
        "5W30 400XSP":            "11",          # Delo 400XSP 5W30
        "5W40 400XSP":            "12",          # Delo 400XSP 5W40
        "Syn-Gear XDM SAE 75W90": "13",          # Delo XDM 75W90
        "Syn-Trans XE SAE 75W90": "15",          # Delo XE 75W90
        "AMT XDT SAE 75W90":      "18",          # Delo AMT XDT (DT12)
        "XV SAE 75W80 (IShift)":  "17",          # Delo XV IShift
        # "Petro Duron UHP 10W30" exists as oil_types code 21 but isn't in Oil Price tab.
    }
    out: dict[str, dict] = {}
    for r in range(1, sh.max_row + 1):
        name = sh.cell(r, 1).value
        if not name:
            continue
        name_s = str(name).strip()
        # Loose match: try exact then substring.
        db_code = None
        for excel_name, code in EXCEL_LABEL_TO_DB_CODE.items():
            if name_s == excel_name or excel_name.lower() in name_s.lower():
                db_code = code
                break
        if not db_code:
            continue
        bulk_sell    = sh.cell(r, 2).value
        bulk_cost    = sh.cell(r, 3).value
        gallon_per_L = sh.cell(r, 6).value  # Gallon Per Litre Cost
        if isinstance(bulk_cost, (int, float)) and bulk_cost > 0:
            out[db_code] = {
                "name": name_s,
                "bulk_cost_per_litre": float(bulk_cost),
                "gallon_per_L": float(gallon_per_L) if isinstance(gallon_per_L, (int, float)) else None,
            }
    return out


# ----------------------------------------------------------------------------
# Parse volume tier brackets from any per-oil GALLON tab header.
# Excel header (15W40 Gallon row 1+2): cols L/M look like "8--20: 10",
# "21--38: 15", "39--46: 20", "47& UP: 25".
# ----------------------------------------------------------------------------
def parse_volume_tiers(wb) -> list[tuple[int, int | None, float]]:
    """Returns (min_litres, max_litres, premium) — applies to all oils."""
    sh = wb["15W40 Gallon"]
    tiers: list[tuple[int, int | None, float]] = []
    for r in range(1, 6):
        bracket = sh.cell(r, 12).value
        premium = sh.cell(r, 13).value
        if not bracket or not isinstance(premium, (int, float)):
            continue
        b = str(bracket).strip()
        # Patterns: "8--20", "21--38", "39--46", "47& UP"
        m = re.match(r"(\d+)\s*[-–—]+\s*(\d+)", b)
        if m:
            tiers.append((int(m.group(1)), int(m.group(2)), float(premium)))
            continue
        m = re.match(r"(\d+)\s*&\s*UP", b, re.I)
        if m:
            tiers.append((int(m.group(1)), None, float(premium)))
            continue
    return tiers


# ----------------------------------------------------------------------------
# Build SQL
# ----------------------------------------------------------------------------
def build_sql(
    filter_data: list[dict],
    db_parts: dict[str, dict],
    db_categories: dict[str, str],
    oil_prices: dict[str, dict],
    tiers: list[tuple[int, int | None, float]],
) -> str:
    lines: list[str] = [
        "-- Generated by scripts/sync-may-2026-costs.py",
        "-- Updates: missing parts, engine_filters wiring for filter-variant engines,",
        "-- oil_types per-litre costs, volume_tiers brackets, and a global $8 labour",
        "-- service_cost attached to each engine's primary oil filter.",
        "-- Idempotent: re-running yields the same end state.",
        "",
        "begin;",
        "",
        "-- =========================================================================",
        "-- 1) Oil-change labour ($8) — a shared service_costs row everyone references",
        "-- =========================================================================",
        "insert into public.service_costs (code, name, cost)",
        "  values ('OIL_CHANGE_LABOUR', 'Oil Change Labour', 8.00)",
        "  on conflict (code) do update set cost = excluded.cost, name = excluded.name;",
        "",
    ]

    # Gather every distinct part_number referenced + whether it's a primary
    # oil filter (col B of Filter Cost) → those parts get the $8 service link.
    distinct_parts: dict[str, dict] = {}
    primary_oil_parts: set[str] = set()
    for f in filter_data:
        # The "primary" oil filter = first entry whose role == "oil"
        primary_set = False
        for ent in f["filters"]:
            pn = ent["part_number"]
            cur = distinct_parts.setdefault(pn, {"cost": ent["cost"], "role": ent["role"]})
            # Keep the highest cost variant if multiple engines reference with diff costs.
            if ent["cost"] > cur["cost"]:
                cur["cost"] = ent["cost"]
            if ent["role"] == "oil" and not primary_set:
                primary_oil_parts.add(pn)
                primary_set = True

    lines += [
        "-- =========================================================================",
        "-- 2) Insert missing parts referenced by the Filter Cost tab",
        "-- =========================================================================",
    ]
    missing_parts = [(pn, info) for pn, info in distinct_parts.items() if pn not in db_parts]
    print(f"Distinct parts referenced: {len(distinct_parts)}; missing in DB: {len(missing_parts)}")
    for pn, info in sorted(missing_parts):
        cat_name = infer_category(pn)
        cat_id = db_categories.get(cat_name.lower())
        if not cat_id:
            # Fall back to "Misc"
            cat_id = db_categories.get("misc")
        if not cat_id:
            print(f"  ! no category_id for {cat_name}; skipping {pn}", file=sys.stderr)
            continue
        # Choose a brand based on prefix
        brand = "OEM"
        if pn.upper().startswith("LF") or pn.upper().startswith("FF") or pn.upper().startswith("FS"):
            brand = "Fleetguard"
        elif pn.upper().startswith("1R") or pn.upper().startswith("AF1"):
            brand = "Cat"
        elif pn.upper().startswith("P5") or pn.upper().startswith("P15") or pn.upper().startswith("P18"):
            brand = "Donaldson"
        elif pn.upper().startswith("B") or pn.upper().startswith("BF"):
            brand = "Baldwin"
        lines.append(
            f"insert into public.parts (part_number, brand, category_id, cost, mhsw_fee, active) "
            f"values ('{sql_escape(pn)}', '{sql_escape(brand)}', '{cat_id}', {info['cost']:.4f}, 0, true) "
            f"on conflict (part_number, brand) do update set cost = excluded.cost;"
        )
    lines.append("")

    lines += [
        "-- =========================================================================",
        "-- 3) Attach the $8 oil-change labour to each primary oil filter part",
        "-- =========================================================================",
        "update public.parts set service_cost_id = (",
        "  select id from public.service_costs where name = 'Oil Change Labour'",
        ")",
        f"where part_number in ({', '.join(repr(pn) for pn in sorted(primary_oil_parts))});",
        "",
    ]

    lines += [
        "-- =========================================================================",
        "-- 4) Wire engine_filters for each filter-variant engine",
        "-- =========================================================================",
    ]
    for f in filter_data:
        dn = f["display_name"]
        for ent in f["filters"]:
            lines.append(
                "insert into public.engine_filters (engine_type_id, part_id, quantity) "
                "select e.id, p.id, 1 "
                "  from public.engine_types e, public.parts p "
                f" where e.display_name = '{sql_escape(dn)}' "
                f"   and p.part_number = '{sql_escape(ent['part_number'])}' "
                "on conflict (engine_type_id, part_id) do update set quantity = excluded.quantity;"
            )
    lines.append("")

    lines += [
        "-- =========================================================================",
        "-- 5) Update oil_types per-litre costs from Excel 'Oil Price' tab",
        "-- =========================================================================",
    ]
    for code, info in oil_prices.items():
        # Gallon per-L cost is stored in DB as gallon_cost_per_litre (per the
        # action layer the value is divided by litres_per_gallon). The DB
        # column name is misleading — it's actually "cost per gallon" in
        # practice. We multiply our per-L back up by litres_per_gallon to
        # match the existing convention.
        bulk = info["bulk_cost_per_litre"]
        sets = [f"bulk_cost_per_litre = {bulk:.4f}"]
        if info["gallon_per_L"] is not None:
            # gallon_cost_per_litre = per_L * litres_per_gallon
            sets.append(
                f"gallon_cost_per_litre = {info['gallon_per_L']:.4f} * coalesce(litres_per_gallon, 3.785)"
            )
        lines.append(
            f"update public.oil_types set {', '.join(sets)} where code = '{sql_escape(code)}';"
        )
    lines.append("")

    lines += [
        "-- =========================================================================",
        "-- 6) Seed volume_tiers (applied to every oil_type) — brackets from Excel headers",
        "-- =========================================================================",
        "-- Tiers from any per-oil GALLON tab header (same brackets across all oils).",
    ]
    for min_l, max_l, premium in tiers:
        # volume_tiers schema: oil_type_id, min_litres, premium. Each oil gets a
        # row per bracket. INSERT for each (oil_type_id, min_litres) pair.
        lines.append(
            f"insert into public.volume_tiers (oil_type_id, min_litres, premium) "
            f"select id, {min_l}, {premium:.2f} from public.oil_types where active = true "
            f"on conflict (oil_type_id, min_litres) do update set premium = excluded.premium;"
        )
    lines.append("")

    lines.append("commit;")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    if not WORKBOOK.exists():
        print(f"workbook not found: {WORKBOOK}", file=sys.stderr)
        return 1

    print(f"Reading {WORKBOOK.relative_to(REPO)} …")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)

    print("Fetching current DB state …")
    db_parts_list = sb("parts?select=id,part_number&limit=5000")
    db_parts = {p["part_number"]: p for p in db_parts_list}
    db_categories_list = sb("part_categories?select=id,name")
    db_categories = {c["name"].lower(): c["id"] for c in db_categories_list}

    print("Parsing Filter Cost tab …")
    filter_data = parse_filter_cost(wb)
    print(f"  → {len(filter_data)} engines with filters")

    print("Parsing Oil Price tab …")
    oil_prices = parse_oil_price(wb)
    print(f"  → {len(oil_prices)} oils with per-litre costs")

    print("Parsing volume tier brackets …")
    tiers = parse_volume_tiers(wb)
    print(f"  → {len(tiers)} brackets: {tiers}")

    print("Generating SQL …")
    sql = build_sql(filter_data, db_parts, db_categories, oil_prices, tiers)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(sql)
    print(f"Wrote {OUT.relative_to(REPO)}")
    print()
    print("Apply with:")
    print(f"  psql \"$DATABASE_URL\" -f {OUT.relative_to(REPO)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
