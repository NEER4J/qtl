#!/usr/bin/env python3
"""
Generate seed SQL to mirror the May 2026 Excel workbook into the platform.

Tabs covered:
  - "15W40" / "15W40 Gallon" and the other per-oil tabs  → engine_sell_prices
  - "Print List"                                          → derived from above (no separate seed)
  - "All Filter Sell Price"                               → parts (override columns added in 0055)

USAGE:
    python3 scripts/sync-may-2026.py            # → writes seed files + report
    python3 scripts/sync-may-2026.py --apply    # → applies directly via Supabase REST

The script reads the current DB to figure out which engines / parts already
exist, then writes:
    supabase/seed/may2026_engine_types.sql           — missing engine_types
    supabase/seed/may2026_engine_sell_prices.sql     — engine_sell_prices upserts
    supabase/seed/may2026_filter_sell_prices.sql     — parts overrides + new parts
    supabase/seed/may2026_REPORT.md                  — dry-run summary

Run the migrations first (0055), then apply the seed files in order:
    psql ... -f supabase/seed/may2026_engine_types.sql
    psql ... -f supabase/seed/may2026_engine_sell_prices.sql
    psql ... -f supabase/seed/may2026_filter_sell_prices.sql
"""
from __future__ import annotations

import json
import os
import re
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install with: pip3 install openpyxl")


REPO = Path(__file__).resolve().parents[1]
WORKBOOK = REPO / "doc" / "May 2026 Standard.xlsx"
SEED = REPO / "supabase" / "seed"

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "https://facbmuiigusuekddtscz.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_KEY:
    sys.exit(
        "SUPABASE_SERVICE_ROLE_KEY env var is required.\n"
        "Get it from .env.local (key: SUPABASE_SERVICE_ROLE_KEY) and prefix the command:\n"
        "  SUPABASE_SERVICE_ROLE_KEY=... python3 scripts/sync-may-2026.py"
    )


# ============================================================================
# OIL TAB → DB OIL CODE MAPPING
# ----------------------------------------------------------------------------
# Picked by comparing Excel "Oil Price" tab cost/L against DB oil_types.
# "15W40" tab uses ~$4.25/L which matches Delo 400 XLE SB ($4.31), not Shell T4
# ($3.84). All other mappings follow the brand name in the tab title.
# ============================================================================
OIL_TAB_MAPPING: list[tuple[str, str, str]] = [
    # (Excel tab name, DB oil_types.code, container)
    ("15W40",             "257004",     "bulk"),    # Delo 400 XLE SB 15W40
    ("15W40 Gallon",      "257004",     "gallon"),
    ("10W30",             "10",         "bulk"),    # Delo 400LE 10W30
    ("10W30 Gallon",      "10",         "gallon"),
    ("T5",                "500010047",  "bulk"),    # Shell T5 10W30
    ("T5 Gallon",         "500010047",  "gallon"),
    ("Petro 10W30",       "21",         "bulk"),    # Petro Duron UHP 10W30
    ("Petro 10W30Gallon", "21",         "gallon"),
    ("T6",                "14",         "bulk"),    # Shell T6 5W30
    ("T6 Gallon",         "14",         "gallon"),
    ("Delo 5W30",         "11",         "bulk"),    # Delo 400XSP 5W30
    ("Delo 5W30 Gallon",  "11",         "gallon"),
    ("Petro 5W30",        "22",         "bulk"),    # Petro 5W30
    ("Petro 5W30 Gallon", "22",         "gallon"),
]


# ============================================================================
# EXCEL ENGINE LABEL → CANONICAL DB display_name
# ----------------------------------------------------------------------------
# The DB previously had collapsed engines (one per make/model) like
# "Cat C7//C10/3126". Excel splits each engine by filter brand
# ("With Cat Filter" / "With Fleetguard Filter"). User chose option A: split
# them in the DB so Excel maps 1:1.
#
# This mapping normalizes Excel typos and inconsistencies, then maps to the
# canonical display_name we want in engine_types. If the canonical name is
# missing, the script will INSERT it.
# ============================================================================
# NOTE: engine_types schema:
#   - display_name is GENERATED ALWAYS AS (manufacturer || ' ' || model) STORED
#   - UNIQUE (manufacturer, model)
#   - oil_capacity_litres NOT NULL CHECK > 0
# So filter-brand variant must live in the `model` column (since manufacturer
# is just the OEM brand), and display_name auto-derives from it.
ENGINE_LABEL_MAPPING: dict[str, tuple[str, str, str]] = {
    # excel label                                  → (manufacturer, model, generated display_name)
    "C 7 / 10 / 3126 With Cat Filter":            ("Cat",        "C7/C10/3126 With Cat Filter",        "Cat C7/C10/3126 With Cat Filter"),
    "C 7 / 10 / 3126 With Fleetguard Filter":     ("Cat",        "C7/C10/3126 With Fleetguard Filter", "Cat C7/C10/3126 With Fleetguard Filter"),
    "C 12 / 3406 With  Cat Filter":               ("Cat",        "C12/3406 With Cat Filter",           "Cat C12/3406 With Cat Filter"),
    "C 12 / 3406 With Cat Filter":                ("Cat",        "C12/3406 With Cat Filter",           "Cat C12/3406 With Cat Filter"),
    "C 12 / 3406 With Fleetguard Filter":         ("Cat",        "C12/3406 With Fleetguard Filter",    "Cat C12/3406 With Fleetguard Filter"),
    "C 13 / 15 With  Cat Filter":                 ("Cat",        "C13/C15 With Cat Filter",            "Cat C13/C15 With Cat Filter"),
    "C 13 / 15 With Cat Filter":                  ("Cat",        "C13/C15 With Cat Filter",            "Cat C13/C15 With Cat Filter"),
    "C 13 / 15 With Fleetguard Filter":           ("Cat",        "C13/C15 With Fleetguard Filter",     "Cat C13/C15 With Fleetguard Filter"),
    "Cummins N14":                                ("Cummins",    "N14",          "Cummins N14"),
    "Cummins 5.9L/6.7L":                          ("Cummins",    "5.9L/6.7L",    "Cummins 5.9L/6.7L"),
    "Cummins ISC / ISL":                          ("Cummins",    "ISC/ISL",      "Cummins ISC/ISL"),
    "Cummins ISB":                                ("Cummins",    "ISB",          "Cummins ISB"),
    "Cummins ISM":                                ("Cummins",    "ISM",          "Cummins ISM"),
    "Cummins ISX":                                ("Cummins",    "ISX/X15",      "Cummins ISX/X15"),
    "Cummins ISX / X15":                          ("Cummins",    "ISX/X15",      "Cummins ISX/X15"),
    "Cummins M11":                                ("Cummins",    "M11",          "Cummins M11"),
    "Detroit With Detroit Filter":                ("Detroit",    "60 Series With Detroit Filter",      "Detroit 60 Series With Detroit Filter"),
    "Detroit With Fleetguard":                    ("Detroit",    "60 Series With Fleetguard Filter",   "Detroit 60 Series With Fleetguard Filter"),
    "Detroit With Detroit Filter (DD8)":          ("Detroit",    "DD8 With Detroit Filter",            "Detroit DD8 With Detroit Filter"),
    "Detroit With Detroit Filter (DD13,DD15,DD16)": ("Detroit",  "DD13/DD15/DD16 With Detroit Filter", "Detroit DD13/DD15/DD16 With Detroit Filter"),
    "Detroit With Fleetguard Filter (DD13,DD15.DD16)": ("Detroit","DD13/DD15/DD16 With Fleetguard Filter","Detroit DD13/DD15/DD16 With Fleetguard Filter"),
    "Duramax / Vortec 8100L":                     ("Duramax",    "/ Vortec 8100L", "Duramax / Vortec 8100L"),
    "F250 / F350 / F550 / F650":                  ("Ford",       "F250/F350/F550/F650", "Ford F250/F350/F550/F650"),
    "Hino":                                       ("Hino",       "With Hino Filter",       "Hino With Hino Filter"),
    "Hino With Fleetguard":                       ("Hino",       "With Fleetguard Filter", "Hino With Fleetguard Filter"),
    "International":                              ("International","With Fleetguard Filter","International With Fleetguard Filter"),
    "International S13":                          ("International","S13",        "International S13"),
    "International Vista":                        ("International","Vista",      "International Vista"),
    "Navistar":                                   ("International","Navistar",   "International Navistar"),
    "Isuzu":                                      ("Isuzu",      "With Isuzu Filter",       "Isuzu With Isuzu Filter"),
    "Isuzu With Fleetguard Filter":               ("Isuzu",      "With Fleetguard Filter",  "Isuzu With Fleetguard Filter"),
    "Mack With Mack Filter":                      ("Mack",       "With Mack Filter",        "Mack With Mack Filter"),
    "Mack With Fleetguard Filter":                ("Mack",       "With Fleetguard Filter",  "Mack With Fleetguard Filter"),
    "Mack With Mack Filter( 52L )":               ("Mack",       "52L With Mack Filter",       "Mack 52L With Mack Filter"),
    "Mack With Fleetguard Filter(52L)":           ("Mack",       "52L With Fleetguard Filter", "Mack 52L With Fleetguard Filter"),
    "MP 7 / MP8 With Mack":                       ("Mack",       "MP 7/8 With Mack Filter",        "Mack MP 7/8 With Mack Filter"),
    "MP 7 / MP8 With Mack Filter":                ("Mack",       "MP 7/8 With Mack Filter",        "Mack MP 7/8 With Mack Filter"),
    "MP 7 / MP8 With Fleetguard":                 ("Mack",       "MP 7/8 With Fleetguard Filter",  "Mack MP 7/8 With Fleetguard Filter"),
    "MP 7 / MP8 With Fleetguard Filter":          ("Mack",       "MP 7/8 With Fleetguard Filter",  "Mack MP 7/8 With Fleetguard Filter"),
    "MP 7/MP8/ MP10 With Mack (52L)":             ("Mack",       "MP 7/8/10 52L With Mack Filter",       "Mack MP 7/8/10 52L With Mack Filter"),
    "MP 7/MP8/ MP10 With Fleetguard ( 52L)":      ("Mack",       "MP 7/8/10 52L With Fleetguard Filter", "Mack MP 7/8/10 52L With Fleetguard Filter"),
    "Max Force 7":                                ("MaxForce",   "7/9/10 With MaxForce Filter",     "MaxForce 7/9/10 With MaxForce Filter"),
    "Max Force 7 / 9 / 10":                       ("MaxForce",   "7/9/10 With MaxForce Filter",     "MaxForce 7/9/10 With MaxForce Filter"),
    "Max Force 7 With Fleetguard":                ("MaxForce",   "7/9/10 With Fleetguard Filter",   "MaxForce 7/9/10 With Fleetguard Filter"),
    "Max Force 7 / 9 / 10 With Fleetguard":       ("MaxForce",   "7/9/10 With Fleetguard Filter",   "MaxForce 7/9/10 With Fleetguard Filter"),
    "Max Force 13":                               ("MaxForce",   "13 With MaxForce Filter",         "MaxForce 13 With MaxForce Filter"),
    "Max Force 13 With Fleetguard":               ("MaxForce",   "13 With Fleetguard Filter",       "MaxForce 13 With Fleetguard Filter"),
    "Mercedes  Benz 4000":                        ("Mercedes",   "Benz 4000 With Mercedes Filter",     "Mercedes Benz 4000 With Mercedes Filter"),
    "Mercedes  Benz 4000 With Fleeguard Filter":  ("Mercedes",   "Benz 4000 With Fleetguard Filter",   "Mercedes Benz 4000 With Fleetguard Filter"),
    "Mercedes  Benz 900":                         ("Mercedes",   "Benz 900 With Mercedes Filter",      "Mercedes Benz 900 With Mercedes Filter"),
    "Mercedes  Benz 900 With Fleetguard Filter":  ("Mercedes",   "Benz 900 With Fleetguard Filter",    "Mercedes Benz 900 With Fleetguard Filter"),
    "Mercedes Benz Sprinter":                     ("Mercedes",   "Benz Sprinter",     "Mercedes Benz Sprinter"),
    "Mitsubushi":                                 ("Mitsubishi", "Engine",   "Mitsubishi Engine"),
    "Mitsubishi":                                 ("Mitsubishi", "Engine",   "Mitsubishi Engine"),
    "Paccer":                                     ("Paccar",     "With Paccar Filter",      "Paccar With Paccar Filter"),
    "Paccer With Fleetguard":                     ("Paccar",     "With Fleetguard Filter",  "Paccar With Fleetguard Filter"),
    "Paccer Small":                               ("Paccar",     "Small", "Paccar Small"),
    "Volvo With Volvo Filter":                    ("Volvo",      "D12 With Volvo Filter",          "Volvo D12 With Volvo Filter"),
    "Volvo With Fleetguard":                      ("Volvo",      "D12 With Fleetguard Filter",     "Volvo D12 With Fleetguard Filter"),
    "Volvo With Volvo Filter (D16 42L)":          ("Volvo",      "D16 42L With Volvo Filter",      "Volvo D16 42L With Volvo Filter"),
    "Volvo With Fleetguard(D16 42L)":             ("Volvo",      "D16 42L With Fleetguard Filter", "Volvo D16 42L With Fleetguard Filter"),
    "Volvo With Volvo Filter (D16 52L)":          ("Volvo",      "D16 52L With Volvo Filter",      "Volvo D16 52L With Volvo Filter"),
    "Volvo With Fleetguard(D16 52L)":             ("Volvo",      "D16 52L With Fleetguard Filter", "Volvo D16 52L With Fleetguard Filter"),
}


# Excel column A values that are NOT real engines (header/reference rows that
# happen to contain numbers in column B). Skip these.
ENGINE_LABEL_BLOCKLIST = {
    "Auto Greaser Refill", "STANDARD JOB LIST", "Jan 1,2026", "Filtre", "Litre",
    "Fuel", "Grease", "Oil", "Oil MHSW", "3  Filter MHSW", "4  Filter MHSW",
    "4.30     Oil", "New 3 Fuel", "selling", "cost",
    # part numbers that slipped into column A
    "21707132", "22480372", "23151592", "23658092", "23920469",
    "FF2200", "FF5776", "FF5825", "FF5971", "LF14000", "QTL1010", "QTL1011",
    "Regular 4 oil Fuel  0372", "Regular 4 oil Fuel  0469",
    "Nanak Shahi ( 15W40 ) Det 60", "J & J Trans ( T6 ) DD15",
    "2071607 Ont ( 15W40 ) ISX",
}


# ============================================================================
# Helpers
# ============================================================================

def sb_get(path: str):
    """GET against Supabase REST."""
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{path}",
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def sql_escape(s: str) -> str:
    return s.replace("'", "''")


def round_99(n: float) -> float:
    """ceil(n) - 0.01 — produces .99 endings to match Excel."""
    import math
    return math.ceil(n) - 0.01


# ============================================================================
# Step 1: collect per-oil sell prices from Excel
# ============================================================================

def collect_engine_oil_prices(wb) -> tuple[list[tuple[str, str, str, float]], dict[str, float]]:
    """Returns (rows, capacities) where rows is (canonical_display_name, oil_code,
    container, sell_price) and capacities is canonical_display_name → litres."""
    rows: list[tuple[str, str, str, float]] = []
    capacities: dict[str, float] = {}
    skipped: dict[str, int] = defaultdict(int)
    for tab, oil_code, container in OIL_TAB_MAPPING:
        if tab not in wb.sheetnames:
            continue
        sh = wb[tab]
        for r in range(6, sh.max_row + 1):
            engine_raw = sh.cell(r, 1).value
            price = sh.cell(r, 2).value
            if not engine_raw or not isinstance(price, (int, float)) or price <= 0:
                continue
            label = str(engine_raw).strip()
            if label in ENGINE_LABEL_BLOCKLIST:
                continue
            if label not in ENGINE_LABEL_MAPPING:
                skipped[label] += 1
                continue
            _, _, canonical = ENGINE_LABEL_MAPPING[label]
            rows.append((canonical, oil_code, container, float(price)))
            # Litre capacity lives in column F (index 6) on every per-oil tab.
            cap = sh.cell(r, 6).value
            if isinstance(cap, (int, float)) and cap > 0 and canonical not in capacities:
                capacities[canonical] = float(cap)
    if skipped:
        print("WARN: skipped Excel engine labels with no mapping:", file=sys.stderr)
        for k, v in sorted(skipped.items()):
            print(f"  '{k}'  ({v} rows)", file=sys.stderr)
    return rows, capacities


# ============================================================================
# Step 2: figure out which canonical engine display_names need to be INSERTed
# ============================================================================

def collect_missing_engines(
    db_engines: list[dict],
    capacities: dict[str, float],
) -> list[tuple[str, str, str, float]]:
    """Returns (manufacturer, model, display_name, oil_capacity_litres) for engines
    NOT in DB. Engines with no matching litre value in Excel are emitted with a
    placeholder 1.0 (must be > 0 to satisfy the check constraint)."""
    have = {e["display_name"].strip().lower() for e in db_engines}
    needed: dict[str, tuple[str, str, str]] = {}
    for label, (mfr, model, dn) in ENGINE_LABEL_MAPPING.items():
        needed[dn] = (mfr, model, dn)
    missing: list[tuple[str, str, str, float]] = []
    for dn, (mfr, model, _) in sorted(needed.items()):
        if dn.strip().lower() not in have:
            litres = capacities.get(dn, 1.0)
            missing.append((mfr, model, dn, litres))
    return missing


# ============================================================================
# Step 3: collect "All Filter Sell Price" rows
# ============================================================================

# Excel column layout (1-indexed):
#  A = Category (e.g. "Air Filter", "Oil Filter")
#  B = Part #
#  C = Without Service raw  | D = Without Service .99
#  E = With Service raw     | F = With Service .99
#  G = Over Counter raw     | H = Over Counter .99
#  I = Void price (unused)  | J = Customer Supplies (global, unused per-part)

def collect_filter_prices(wb) -> list[dict]:
    sh = wb["All Filter Sell Price"]
    out: list[dict] = []

    def clean(v):
        """Return v as float when it's a positive number, else None. Excel has
        a few bogus negative values (formula errors / typos) that violate the
        parts.*_price >= 0 check constraint — drop those to NULL."""
        if not isinstance(v, (int, float)):
            return None
        f = float(v)
        return f if f >= 0 else None

    for r in range(3, sh.max_row + 1):
        cat = sh.cell(r, 1).value
        pn = sh.cell(r, 2).value
        if not cat or not pn:
            continue
        cat_s = str(cat).strip()
        pn_s = str(pn).strip()
        if not cat_s or not pn_s or cat_s.lower() in {"manufactur", "manufacturer"}:
            continue
        # The three sell prices (use the .99-rounded columns: D, F, H)
        without = clean(sh.cell(r, 4).value)
        with_   = clean(sh.cell(r, 6).value)
        over    = clean(sh.cell(r, 8).value)
        # only emit rows that have at least one price
        if all(v is None for v in (without, with_, over)):
            continue
        out.append({
            "category": cat_s,
            "part_number": pn_s,
            "without_service": without,
            "with_service":    with_,
            "over_counter":    over,
        })
    return out


# ============================================================================
# SQL emitters
# ============================================================================

def emit_engine_types_sql(missing: list[tuple[str, str, str, float]]) -> str:
    lines = [
        "-- Generated by scripts/sync-may-2026.py",
        "-- Adds engine_types entries missing from the platform but present in the May 2026 Excel.",
        "-- Existing engines are left untouched. display_name is a GENERATED column so we",
        "-- only insert manufacturer + model.",
        "",
        "begin;",
        "",
    ]
    for mfr, model, dn, litres in missing:
        lines.append("-- " + dn)
        lines.append("insert into public.engine_types (manufacturer, model, oil_capacity_litres)")
        lines.append(f"  values ('{sql_escape(mfr)}', '{sql_escape(model)}', {litres:.2f})")
        lines.append("  on conflict (manufacturer, model) do nothing;")
        lines.append("")
    lines.append("commit;")
    return "\n".join(lines) + "\n"


def emit_engine_sell_prices_sql(prices: list[tuple[str, str, str, float]]) -> str:
    lines = [
        "-- Generated by scripts/sync-may-2026.py",
        "-- Upserts engine_sell_prices from the per-oil tabs in the May 2026 Excel.",
        "-- Looks up engine by display_name and oil by code. Apply AFTER",
        "-- supabase/seed/may2026_engine_types.sql.",
        "",
        "begin;",
        "",
    ]
    for canonical, oil_code, container, price in prices:
        lines.append(
            "insert into public.engine_sell_prices (engine_type_id, oil_type_id, container, sell_price)"
        )
        lines.append(f"select e.id, o.id, '{container}', {price:.2f}")
        lines.append("  from public.engine_types e, public.oil_types o")
        lines.append(f" where e.display_name = '{sql_escape(canonical)}'")
        lines.append(f"   and o.code = '{sql_escape(oil_code)}'")
        lines.append(
            "  on conflict (engine_type_id, oil_type_id, container) "
            "do update set sell_price = excluded.sell_price;"
        )
        lines.append("")
    lines.append("commit;")
    return "\n".join(lines) + "\n"


def emit_filter_prices_sql(filters: list[dict], db_parts: dict[str, dict], db_categories: dict[str, str]) -> tuple[str, list[dict]]:
    """Returns (sql, missing_parts_list)."""
    lines = [
        "-- Generated by scripts/sync-may-2026.py",
        "-- Mirrors the 'All Filter Sell Price' Excel tab into parts.* override columns",
        "-- (parts.without_service_price, parts.with_service_price, parts.over_counter_price).",
        "-- Apply AFTER migration 0055_parts_sell_price_overrides.sql.",
        "--",
        "-- Only updates existing parts (matched by part_number). New parts must be created",
        "-- via the Pricing catalogue admin or a separate seed (see may2026_REPORT.md).",
        "",
        "begin;",
        "",
    ]
    matched = 0
    missing: list[dict] = []
    for f in filters:
        pn = f["part_number"]
        db = db_parts.get(pn) or db_parts.get(pn.lower())
        if not db:
            missing.append(f)
            continue
        matched += 1
        sets = []
        if f["without_service"] is not None:
            sets.append(f"without_service_price = {f['without_service']:.2f}")
        else:
            sets.append("without_service_price = null")
        if f["with_service"] is not None:
            sets.append(f"with_service_price    = {f['with_service']:.2f}")
        else:
            sets.append("with_service_price    = null")
        if f["over_counter"] is not None:
            sets.append(f"over_counter_price    = {f['over_counter']:.2f}")
        else:
            sets.append("over_counter_price    = null")
        lines.append(f"update public.parts set {', '.join(sets)}")
        lines.append(f"  where part_number = '{sql_escape(pn)}';")
        lines.append("")
    lines.append("commit;")
    lines.append("")
    lines.append(f"-- Matched {matched} parts. {len(missing)} Excel rows had no matching part_number in DB (see report).")
    return "\n".join(lines) + "\n", missing


# ============================================================================
# Report
# ============================================================================

def emit_report(
    prices: list,
    missing_engines: list,
    filters: list,
    missing_parts: list,
    db_engines: list,
    db_parts: dict,
) -> str:
    label_to_canonical: dict[str, str] = {l: c[2] for l, c in ENGINE_LABEL_MAPPING.items()}
    by_canon: dict[str, list[str]] = defaultdict(list)
    for label, canon in label_to_canonical.items():
        by_canon[canon].append(label)
    # also tally engine_sell_prices rows per canonical
    sell_per_canon: dict[str, int] = defaultdict(int)
    for canonical, _, _, _ in prices:
        sell_per_canon[canonical] += 1

    matched_filters = len(filters) - len(missing_parts)
    cats: dict[str, int] = defaultdict(int)
    for f in missing_parts:
        cats[f["category"]] += 1

    lines = [
        "# May 2026 sync — dry-run report",
        "",
        "Generated by `scripts/sync-may-2026.py` from `doc/May 2026 Standard.xlsx`.",
        "",
        "## Engines",
        "",
        f"- Excel labels mapped: {len(ENGINE_LABEL_MAPPING)} (deduped to {len(by_canon)} canonical engine names)",
        f"- Already in DB engine_types: {len(by_canon) - len(missing_engines)}",
        f"- **Will be INSERTED** ({len(missing_engines)}):",
    ]
    for mfr, model, dn, litres in missing_engines:
        lines.append(f"  - {dn}  *(manufacturer={mfr}, model={model}, litres={litres})*")
    lines += [
        "",
        "## engine_sell_prices",
        "",
        f"- Total upserts: **{len(prices)}** (one per engine × oil × container)",
        f"- Currently in DB: 0",
        f"- Oil tab → DB oil mapping:",
    ]
    for tab, code, cont in OIL_TAB_MAPPING:
        lines.append(f"  - `{tab}` → oil_types.code=`{code}`, container=`{cont}`")
    lines += [
        "",
        "## All Filter Sell Price (parts overrides)",
        "",
        f"- Excel rows with at least one sell price: **{len(filters)}**",
        f"- Matched to existing parts (will UPDATE override columns): **{matched_filters}**",
        f"- **No matching part_number in DB** (will NOT be inserted by seed): **{len(missing_parts)}**",
        "",
        "  Breakdown of unmatched by Excel category:",
    ]
    for cat, n in sorted(cats.items(), key=lambda kv: -kv[1]):
        lines.append(f"  - {cat}: {n}")
    lines += [
        "",
        "## Next steps",
        "",
        "1. Apply migration 0055 (`supabase db push` or run via SQL editor).",
        "2. Apply seed in order:",
        "   ```",
        "   psql \"$DATABASE_URL\" -f supabase/seed/may2026_engine_types.sql",
        "   psql \"$DATABASE_URL\" -f supabase/seed/may2026_engine_sell_prices.sql",
        "   psql \"$DATABASE_URL\" -f supabase/seed/may2026_filter_sell_prices.sql",
        "   ```",
        "3. Update `getAllFilterSellPrices` to read the new override columns (see TODO in lib/actions/pricing.ts).",
        "4. Visit `/pricing/oil-detail/257004?container=bulk` (15W40), `/pricing/print-list`, and `/pricing/all-filter-price` to verify.",
        "",
    ]
    return "\n".join(lines) + "\n"


# ============================================================================
# Main
# ============================================================================

def main() -> int:
    if not WORKBOOK.exists():
        print(f"workbook not found: {WORKBOOK}", file=sys.stderr)
        return 1

    print(f"Reading {WORKBOOK.relative_to(REPO)} …")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)

    print("Fetching current DB state …")
    db_engines = sb_get("engine_types?select=display_name,manufacturer,model&active=eq.true&limit=500")
    db_parts_list = sb_get("parts?select=id,part_number,brand,description,cost,mhsw_fee&limit=5000")
    db_parts: dict[str, dict] = {p["part_number"]: p for p in db_parts_list}
    db_categories_list = sb_get("part_categories?select=id,name")
    db_categories: dict[str, str] = {c["name"].lower(): c["id"] for c in db_categories_list}

    print("Collecting engine sell prices from Excel …")
    prices, capacities = collect_engine_oil_prices(wb)
    print(f"  → {len(prices)} (engine, oil, container) rows")
    print(f"  → {len(capacities)} engines have litre capacity in Excel")

    missing_engines = collect_missing_engines(db_engines, capacities)
    print(f"  → {len(missing_engines)} engine_types to INSERT")

    print("Collecting filter sell prices from Excel …")
    filters = collect_filter_prices(wb)
    print(f"  → {len(filters)} filter rows with at least one price")

    print("Generating SQL files …")
    SEED.mkdir(parents=True, exist_ok=True)
    (SEED / "may2026_engine_types.sql").write_text(emit_engine_types_sql(missing_engines))
    (SEED / "may2026_engine_sell_prices.sql").write_text(emit_engine_sell_prices_sql(prices))
    filter_sql, missing_parts = emit_filter_prices_sql(filters, db_parts, db_categories)
    (SEED / "may2026_filter_sell_prices.sql").write_text(filter_sql)
    (SEED / "may2026_REPORT.md").write_text(
        emit_report(prices, missing_engines, filters, missing_parts, db_engines, db_parts)
    )
    print(f"  → {SEED}/may2026_*.sql + may2026_REPORT.md")
    print()
    print(f"Matched parts: {len(filters) - len(missing_parts)} / {len(filters)}")
    print(f"Unmatched Excel filter rows (no DB part_number): {len(missing_parts)}")
    print()
    print("Review supabase/seed/may2026_REPORT.md before applying.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
